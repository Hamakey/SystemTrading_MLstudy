class Loop_def():
    
    def __init__(self):
        print("Loop_class입니다.")
        
    
    import pandas as pd
    import openpyxl
    from openpyxl import Workbook
    import basicInfo_list as b_info
    
    def crawling_loop(target_data_def, request_def, path, load_sheet_name, save_sheet_name):
    
            import pandas as pd
            import openpyxl
            from openpyxl import Workbook
            import basicInfo_list as b_info
            #############중복 openpyxl.load_workbook()을 방지하기위한 변수 global화##########
            global code_data
            global thefile
            global create_ws
            global code_data_length
            ###############################################################################
            code_data=pd.read_excel(path,load_sheet_name)
            code_data_length=len(code_data.index)
            
            thefile=openpyxl.load_workbook(path,data_only=True)
            create_ws=thefile.create_sheet(save_sheet_name)
            category, length = b_info.basicInfo_list(1)
            target=target_data_def
            request=request_def
            for i in range(code_data_length):
                global code_row
                code_row=i
                target_code=code_data.loc[i][1]   ##종목코드 
                print("data_gathering (%s/%s) 진행중" %(i,i/code_data_length))
                target
                request
                


    def trdata_loop(self,sTrCode, sRQName, code_row, save_file_name):
        import pandas as pd
        import openpyxl
        from openpyxl import Workbook
        import basicInfo_list as b_info
        category, length = b_info.basicInfo_list(1)
        for i in range(length):
            
            category, length = b_info.basicInfo_list(i)
            data_set=self.dynamicCall("GetCommData(String, String, String, String)", sTrCode, sRQName, 0, category)
            data_set=data_set.strip()
            write_data1=create_ws.cell(code_row+2, 1).value=code_data.iloc[code_row][1]
            write_data2=create_ws.cell(row=1,column=i+2,value=category)
            write_data3=create_ws.cell(row=code_row+2,column=i+2,value=data_set)
            print("i값=%s, code_row값=%s 에서 %s = %s" %(i+2,code_row+2, category ,data_set))
            self.search_basicInfo_event_loop.exit()
            save_ws=thefile.save(save_file_name)
            