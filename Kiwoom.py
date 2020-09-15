##############[Check] 효율적으로 코드를 줄일 수 있을지 고민필요
##############Update 이력 재정리해야함
##############추가 과제
#단기목표(21년 06월)_기본적인 Trade, Backtesting을 위한 Data Crawling
#1. 재무제표에서 지표간의 관계 재정리(M/L을 위한 피쳐 선정목적)
#2. 선정 지표 10종류까지 늘리기
#3. 매매/체결 Logic 구현
#4. 재무제표의 분기/연간 지표에 따른 Data 시각화(시계열 기준)
#중기목표(21년 3월)_Back testing 가능하도록 Logic 구현! 
#1. [M/L]선정 지표를 이용하여 M/L 적용(회귀(?), 어떤방식이 적절할지 공부필요)
#2. [M/L]시계열 데이터 결측치 처리 방식 공부해야할 것(가격 예측의 가장 큰 Noise)
#3. [M/L]가격결정 모델 Update
#4. [Trading]매도 Sign, 매수 Sign 결정 모델 추가
#5. 증권사 Quant의 역할 공부(Quant의 의사결정 방법 Logic화 하여 알고리즘 구현)
#6. 주가예측 관련 논문 1편 선정/공부
#장기 목표()
#1. [자연어처리(?) 공부해야함]뉴스기사 제목 크롤링 통해서 긍정/부정 구분하는 D/L 적용(장기목표)
#2.  UI 구현
import os
import sys
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtTest import *
from PyQt5.QAxContainer import *
from errorCode import *
import pandas as pd
from time import *
import openpyxl
from openpyxl import Workbook
import basicInfo_list as b_info
class Kiwoom(QAxWidget):
    
    def __init__(self):
        super().__init__()
    
        print("kiwoom class 입니다.")
        #########eventloop 모듈######### 하나로 줄여서 사용할 수 없을지 테스트 필요
        #########QEventLoop로 쓰레드 관리하는 방법 다시 공부해야함#################
        self.login_event_loop = None
        self.detail_account_info_event_loop = QEventLoop()
        self.search_basicInfo_event_loop = None
        self.detail_account_mystock = None
        self.detail_account_mystock = QEventLoop()
        self.calculator_event_loop = QEventLoop()
        #########
        
        #########스크린번호 모음
        
        self.screen_my_info = '2000' #계좌 관련한 스크린 번호
        self.screen_calculation_stock = "4000" #계산용 스크린 번호
        self.screen_real_stock = "5000" #종목별 할당할 스크린 번호
        self.screen_meme_stock = "6000" #종목별 할당할 주문용스크린 번호
        self.screen_start_stop_real = "1000" #장 시작/종료 실시간 스크린번호
        ########################################
        #########
        
        #########변수모음
        self.account_num=None
        # self.insert_company_code()
        
        ######### 계좌 관련 변수
        self.use_money = 0
        self.use_money_percent = 0.5
        ######### 변수 모음
        self.account_stock_dict = {}
        
        ######### 종목분석용 데이터 모음
        self.calcul_data = []
        
        ######### 초기 셋팅 함수들 바로 실행
        self.get_ocx_instance() #OCX 방식을 파이썬에 사용할 수 있게 변환해 주는 함수
        self.event_slots()
        self.signal_login_commConnect()
        self.get_account_info() # 기본 계좌정보
        self.detail_account_info()#예수금 가져오기, 여기서 계좌번호/비밀번호/비번매체구분/조회구분 입력해서 self로 올려뒀음.-> 다음 조회시 바로 입력받을 수 있음.
        self.search_basicInfo() # 종목별 기본정보 _PER, ROE, PBR, 당기순이익 등..
        self.detail_account_info() # 계좌상세(보유 현황 등)
        self.not_concluded_account() # 체결현황 요청 코드
        
        self.calculator_fnc() # 종목분석코드
        #########event loop 모음

        #########
        
    def get_ocx_instance(self):
        self.setControl("KHOPENAPI.KHOpenAPICtrl.1")
        ##키움 windows registry 접속
        
    def event_slots(self):
        self.OnEventConnect.connect(self.login_slot) # 로그인 관련 이벤트
        self.OnReceiveTrData.connect(self.trdata_slot)
        
        
    def login_slot(self,errCode):
        
        print(errors(errCode))     ###각종 에러 별 메시지 발생
        
        
        #로그인 처리가 완료됐으면 이벤트 루프를 종료한다.
        self.login_event_loop.exit()
    
    def signal_login_commConnect(self):
        self.dynamicCall("CommConnect")
        
        self.login_event_loop = QEventLoop()  
        # 동시동작을 통한 쓰레드에러(현재코드 진행중 다음코드를 진행하다가 에러발생하는 경우)를 
        # 방지하기위한 block(현재 진행중인 code를 완료하기전까지는 다음 코드를 실행시키지 않음.)
        # 을 포함하고 있어서 Tread(≒연산 core) 충돌을 방지함.
        self.login_event_loop.exec_()
        
    def get_account_info(self):
        # KOA 양식이 1:1이므로 "ACCNO", "USER_ID"로 2개의 argument를 요청하면 안됨.
        # user_id = self.dynamicCall("GetLogininfo(QString)","USER_ID")   user id 받아오는 예시
        account_list = self.dynamicCall("GetLoginInfo(QString)", "ACCNO") # 계좌번호 반환
        account_num = account_list.split(';')[0]
        self.account_num = account_num
        print("나의 계좌번호 : %s" % account_num)
        
    def detail_account_info(self):
        
        self.dynamicCall("SetInputValue(QString, QString)", "계좌번호",self.account_num)
        self.dynamicCall("SetInputValue(QString, QString)", "비밀번호","0000") #
        self.dynamicCall("SetInputValue(QString, QString)", "비밀번호입력매체구분","00") #비밀번호입력매체구분 = 00
        self.dynamicCall("SetInputValue(QString, QString)", "조회구분","2")  #조회구분 = 1:추정조회, 2:일반조회
        self.dynamicCall("CommRqData(QString, QString, QString, QString)","예수금상세현황요청","opw00001", "0","2000")# CommRqData( "RQName"	,  "opw00001"	,  "0"	,  "화면번호"); 
        #########tr data 요청까지 완료
        self.detail_account_info_event_loop = QEventLoop()##이벤트루프 진입 코드
        self.detail_account_info_event_loop.exec_()##이벤트루프 시작 코드
        ######### EVENT LOOP라는 곳으로 들어와서 병렬처리 진행되는 부분임 ---> 추가공부필요
    
    def search_basicInfo(self):
        
        import pandas as pd
        ####################Ver.02 Update 내용#########################################
        #############중복 openpyxl.load_workbook()을 방지하기위한 변수 global화##########
        #############Self 와 Global의 차이점 공부 필요함. 아직 잘 모르고 사용하고있음######
        global path
        global code_data
        global thefile
        global create_ws
        global code_data_length
        global target_code_list
        ###############################################################################
        #############이부분의 파일 불러오고 Data 받는 것도 Code Review 필요##############
        path=r'C:\Users\user\Desktop\Python\quant_trading\Kiwoon\extracted_data_stocks.xlsx'
        code_data=pd.read_excel(path,'소프트웨어 개발 및 공급업')
        code_data_length=len(code_data.index)
        sheet_check=openpyxl.load_workbook(path)
        count_sheet_Qty=len(sheet_check.sheetnames)
        data_sheet_name='소프트웨어 개발 및 공급업 조회'
        print(sheet_check.sheetnames)

        
        thefile=openpyxl.load_workbook(path,data_only=True)
        create_ws=thefile.create_sheet(data_sheet_name)
        category, length = b_info.basicInfo_list(1)
        target_code_list = [0 for i in range(code_data_length)]
        for i in range(code_data_length):
        
        #####################기본정보조회 Pass data 갱신여부까지는 확인불가####################
            for j in range(count_sheet_Qty) :
                check_result=sheet_check.sheetnames[j]
                if data_sheet_name == sheet_check.sheetnames[j] :
                    break
                
            if check_result == data_sheet_name:
                print("기본 data있음, 기본정보조회 pass")
                break
        ####################################################################################
            global code_row
            code_row=i
            target_code=code_data.loc[i][1]   ##종목코드 반복입력
            target_code_list[i] = int(target_code.lstrip("A"))
            self.target_code_list= target_code_list
            
            print("i번쨰 search_basicInfo 함수의 값=%s"%target_code)
            self.dynamicCall("SetInputValue(QString,QString)", "종목코드", target_code)
            self.dynamicCall("CommRqData(QString, QString, QString, QString)", "주식기본정보요청", "opt10001", "0", self.screen_my_info)
            ######### tr data 요청까지 완료
            self.search_basicInfo_event_loop = QEventLoop()
            self.search_basicInfo_event_loop.exec_()
            
            if len(code_data.index)<=3600*5:
                
                sleep(0.21) #### 키움api 1s당 5회 조회, 1시간당 1000회 조회 제한 대비
            else :
                print("조회제한 초과")
                break
            
            
    def detail_account_mystock(self, sPrevNext="0"):
        ####sPrevNext 를 0으로하면 single data만 받아올 수 있음, multi data를 받아오려면 다르게 설정해야함.
        print("계좌평가잔고내역요청")
        self.dynamicCall("SetInputValue(QString, QString)", "계좌번호",self.account_num)
        self.dynamicCall("SetInputValue(QString, QString)", "비밀번호","0000") #
        self.dynamicCall("SetInputValue(QString, QString)", "비밀번호입력매체구분","00") #비밀번호입력매체구분 = 00
        self.dynamicCall("SetInputValue(QString, QString)", "조회구분","2")  #조회구분 = 1:추정조회, 2:일반조회
        self.dynamicCall("CommRqData(QString, QString, QString, QString)","계좌평가잔고내역요청","opw00018", sPrevNext, self.screen_my_info)# CommRqData( "RQName"	,  "opw00001"	,  "0"	,  "화면번호"); 
        #########tr data 요청까지 완료
        
        self.detail_account_mystock.exec()
        
    def not_concluded_account(self, sPrevNext="0"):
        
        print("실시간 미수체결현황조회")
        
        self.dynamicCall("SetInputValue(QString, QString)", "계좌번호",self.account_num)
        self.dynamicCall("SetInputValue(QString, QString)", "전체종목구분","0") #전체종목구분 = 0:전체, 1:종목
        self.dynamicCall("SetInputValue(QString, QString)", "체결구분","0") #체결구분 = 0:전체, 2:체결, 1:미체결
        self.dynamicCall("SetInputValue(QString, QString)", "매매구분","0")  #매매구분 = 0:전체, 1:매도, 2:매수
        self.dynamicCall("CommRqData(QString, QString, int, QString)","실시간체결내역요청","opt10075", sPrevNext, self.screen_my_info)# CommRqData( "RQName"	,  "opw00001"	,  "0"	,  "화면번호"); 
        
        self.detail_account_mystock.exec()
        
    def get_code_list_by_market(self, market_code):
        
        
        code_list = self.dynamicCall("GetCodeListByMarket(QString)", market_code)
        code_list = code_list.split(";")[:-1]
        return code_list
    
    def calculator_fnc(self):
        ###############실전 code######################
        code_list = self.get_code_list_by_market("10")
        print("코스닥 갯수 %s" % len(code_list))
        #############################################
        
        ########연습을위한 숫자제한code########
        # if idx==11:
        #     pass
        # else:
        ################################
        # ##################실전code############
        for idx, code in enumerate(code_list):
            self.dynamicCall("DisconnectRealData(QSting)", self.screen_calculation_stock)
            print("%s / %s 코스닥 종목코드 : %s 업데이트 중" % (idx+1, len(code_list), code))
        
            self.day_kiwoom_db(code=code)
            if idx==11:
                break
            
            # #####################################
            
        # print("코스닥 갯수 %s" % len(code_data))
        
        # for idx,code in enumerate(code_data):
        #     self.dynamicCall("DisconnectRealData(QSting)", self.screen_calculation_stock)
        #     print("%s / %s 코스닥 종목코드 : %s 업데이트 중" % (idx+1, len(code_data), code))
            
        #     self.day_kiwoom_db(code=code)

    def day_kiwoom_db(self, code=None, date=None, sPrevNext='0'):
        QTest.qWait(3600) # 동시성처리에 대해 방지하는 delay code//모듈 내용 재확인 필요
        self.dynamicCall("SetInputValue(QString, QString)", "종목코드", code)
        self.dynamicCall("SetInputValue(QString, QString)", "수정주가구분", "1")
        
        if date != None:
            self.dynamicCall("SetInputValue(QString, QString)", "기준일자", date)
            
        self.dynamicCall("CommRqData(QString, QString, int, QString)", "주식일봉차트조회", "opt10081", sPrevNext, self.screen_calculation_stock)  # Tr서버로 전송 -Transaction
        
        self.calculator_event_loop.exec_() 
        # 1개 종목 조회중인데 바로 데이터조회 완료되기전에 다른종목을 동시에 시작하려는걸 막아주기위한
        # event_loop임
        # 동시성 처리를 지원하기때문에 위에 print의 code_list에 1 by 1 으로 code를 넘기는게 아니라 하나 넘기는 동시에 새로운게 가지고있는 상태에서 들어옴
    
    def trdata_slot(self, sScrNo, sRQName, sTrCode, sRecordName, sPrevNext):
        '''
        TR요청을 받는 구역, SLOT 임.
        Parameters
        ----------
        sScrNo :  스크린 번호
        sRQName :  내가 요청했을때 지은 이름
        sTrCode :  요청ID, RT코드
        sRecordName :  사용 안함
        sPrevNext :  다음페이지가 있는지 확인
        Returns
        ----------
        '''
        if sRQName =='예수금상세현황요청':
            deposit=self.dynamicCall("GetCommData(QString, QString, QString, QString)", sTrCode, sRQName, 0, "예수금")
            self.deposit=deposit
            print("예수금 형변환 완료 : %s" % int(deposit))
            
            ######구매금액 기준정립 -> 구매마지노선 선정 알고리즘 작성필요
            print("구매금액 계산 시작")
            self.use_money = int (deposit) * self.use_money_percent
            self.use_money = self.use_money / 4
            print("구매완료 ->구매금액 형변환 완료 : %s" % float(self.use_money))
            ######
            
            ok_deposit=self.dynamicCall("GetCommData(QString, QString, QString, QString)", sTrCode, sRQName, 0, "출금가능금액")
            self.ok_deposit=ok_deposit
            print("출금가능금액 형변환 완료 : %s" % int(ok_deposit))
            
            self.detail_account_info_event_loop.exit()
            
        if sRQName == '주식기본정보요청' :
            ##################read_excel을 search_basicInfo로 두번해서 저장이안됐던거임################
            # path=r'C:\Users\user\Desktop\Python\quant_trading\Kiwoon\data3.xlsx'
            # code_data=pd.read_excel(path,'소프트웨어 개발 및 공급업')
            # code_data_length=len(code_data.index)
            # thefile=openpyxl.load_workbook(path,data_only=True)
            # create_ws=thefile.create_sheet('소프트웨어 개발 및 공급업 data')
            ##############################################################################
            category, length = b_info.basicInfo_list(1)
            
            
            for i in range(length):###enumerate 로 돌리면 되지않나??? idx랑 값 둘다 동시에 주는게 enumerate이니까
                
                category, length = b_info.basicInfo_list(i)
                types = b_info.data_type(i)
                write_data1=create_ws.cell(code_row+2, 1).value=code_data.iloc[code_row][1]
                write_data2=create_ws.cell(row=1,column=i+2,value=category)
                data_set=self.dynamicCall("GetCommData(QString, QString, QString, QString)", sTrCode, sRQName, 0, category)
                
                data_set=data_set.strip()
                print("i값=%s, code_row값=%s data set=%s" %(i+2,code_row+2,data_set))
                write_data3=create_ws.cell(row=code_row+2,column=i+2,value=data_set)
                ##################Ver
                ##################형변환 logic, Data_set int 변환 실패해결해야함################
                # if types =="intiger":
                #     data_set=data_set.strip()
                #     print("strip 이후의 data set은 %s" %data_set)
                #     data_set=int(data_set)
                #     write_data3=create_ws.cell(row=j+2,column=i+2,value=data_set)
                # elif types =="float":
                #     data_set=data_set.strip()
                #     print("strip 이후의 data set은 %s" %data_set)
                #     data_set=float(data_set)
                #     write_data3=create_ws.cell(row=j+2,column=i+2,value=data_set)
                # else :
                #     data_set=data_set.strip()
                #     print("strip 이후의 data set은 %s" %data_set)
                #     write_data3=create_ws.cell(row=j+2,column=i+2,value=data_set)
                ##############################################################################
                print("총 %s개의 data_set에서 %s = %s" %(length ,category ,data_set))
                
                self.search_basicInfo_event_loop.exit()
                # save_ws=thefile.save('data4.xlsx')
            save_ws=thefile.save('extracted_data_stocks.xlsx')
            
            
        if sRQName =='계좌평가잔고내역요청':
            total_buy_money=self.dynamicCall("GetCommData(QString, QString, QString, QString)", sTrCode, sRQName, 0, "총매입금액")
            total_buy_money_result = int(total_buy_money)
            print("총 매입금액 %s" % total_buy_money_result)
            total_profit_loss_rate=self.dynamicCall("GetCommData(QString, QString, QString, QString)", sTrCode, sRQName, 0, "총수익률(%)")
            ######위의 0 은 No(보유종목 순서를 매기는 번호 0번종목, 1번종목, 2번종목 등등 n-1번종목까지 총 n개의 종목을 의미함 = 종목의 위치값)
            total_profit_loss_rate_result = float(total_profit_loss_rate)
            print("총 수익률(%s) : %s" %("%", total_buy_profit_loss_rate))
            
            mystocks_count = self.dynamicCall("GetRepeatCnt(QSting, QString)", sTrCode, sRQName)
            ########GetRepeatCnt 는 multi data 조회용 명령어이며, 한번에 20개까지밖에 조회하지 못해서 sPrevNext에 조건문 걸어서 다음거로 넘어가야함
            for i in range(mystocks_count):
                ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                code = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "종목코드")
                code_name = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "종목명")
                stock_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "보유수량")  # 보유수량 : 000000000000010
                buy_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "매입가")  # 매입가 : 000000000054100
                profit_rate = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "수익률(%)")  # 수익률 : -000000001.94
                current_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "현재가")  # 현재가 : 000000003450
                total_signed_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "매입금액")
                possible_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "매매가능수량")
                
                if code in self.account_stock_dict :
                    pass
                else:
                    self.account_stock_dict.update({code:{}})
                
                code = code.strip()[1:]
                code_name = code_name.strip()
                stock_quantity = int(stock_quantity.strip())
                buy_price = int(buy_price.strip())
                profit_rate = float(profit_rate.strip())
                current_price = int(current_price.strip())
                total_signed_price = int(total_signed_price.strip())
                possible_quantity = int(possible_quantity.strip())
                
                print("내 보유주식 정보 Loading 중")
                print("종목code : %s \n 종목명 : %s " % (code, code_name))
                print("보유수량 : %s \n 수익률 : %s \n 매입가 : %s \n현재가 : %s \n " %(stock_quantity, profit_rate , buy_price, current_price))
                print("총매입금액 : %s \n 총매매가능금액 : %s  " %(possible_quantity, total_signed_price))
                ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                self.account_stock_dict[code].update({"종목명" : code_nm})
                self.account_stock_dict[code].update({"보유수량" : stock_quantity})
                self.account_stock_dict[code].update({"수익률" : profit_rate})
                self.account_stock_dict[code].update({"매입가" : buy_price})
                self.account_stock_dict[code].update({"n현재가" : current_price})
                self.account_stock_dict[code].update({"총매입금액" : total_signed_price})
                self.account_stock_dict[code].update({"총매매가능금액" : possible_quantity})
                
                print("계좌에 보유한 총 종목 수 %s" % len(self.account_stock_dict))
                
                if sPrevNext =="2":
                    self.detail_account_mystock(sPrevNext="2")
                else:
                    self.detail_account_mystock.exit()
        elif sRQName =="실시간체결내역요청":
            
            mystocks_count = self.dynamicCall("GetRepeatCnt(QSting, QString)", sTrCode, sRQName)
            for i in range(mystocks_count):
                ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                code = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "종목코드")
                code_name = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "종목명")
                order_num = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "주문번호")
                order_status = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "주문상태")
                order_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "주문수량")
                order_price = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "주문가격")
                order_type = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "주문구분")
                order_unsigned_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "미체결수량")
                order_signed_quantity = self.dynamicCall("GetCommData(QString, QString, int, QString", sTrCode, sRQName, i, "체결수량")
                
                ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                code = code.strip()
                code_name = code_name.strip()
                order_num = int(order_num.strip())
                order_status = order_status.strip()
                order_quantity = int(order_quantity.strip())
                order_price = int(order_price.strip())
                order_type = order_type.strip().lstrip('+').lstrip("-")
                order_unsigned_quantity = int(order_unsigned_quantity.strip())
                order_signed_quantity = int(order_signed_quantity.strip())
                
                if order_num in self.not_account_stock_dict :
                    pass
                else:
                    self.not_account_stock_dict[order_num] = {}
                    ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                nasd = self.not_account_stock_dict[order_num]
                nasd.update({"종목코드": code})
                nasd.update({"종목명": code_name})
                nasd.update({"주문번호": order_num})
                nasd.update({"주문상태": order_status})
                nasd.update({"주문수량": order_quantity})
                nasd.update({"주문가격": order_price})
                nasd.update({"주문구분": order_type})
                nasd.update({"미체결수량": order_unsigned_quantity})
                nasd.update({"체결수량": order_signed_quantity})
                print("체결진행중인 종목 : %s " % self.not_account_stock_dict[order_num])
                
            self.detail_account_mystock.exit()
        elif "주식일봉차트조회" == sRQName:
            
            code = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, 0, "종목코드") # single data 라 value3이 0으로 드감 아니며 for로 i 돌려야함
            code = code.strip()
            print('주식일봉차트조회요청 중 : %s ' % code)
            # 아래의 for 문 대신에 GetCommDataEx 함수를 쓰면 그아래와같이 600일치 데이터를 한 번에 받아올 수 있음.
            # 나중에 변경하여 Code 간소화 시킬 것.
            # data = self.dynamicCall("GetCommDataEx(QString, QString)", sTrCode, sRQName)
            # [[‘’, ‘현재가’, ‘거래량’, ‘거래대금’, ‘날짜’, ‘시가’, ‘고가’, ‘저가’. ‘’], [‘’, ‘현재가’, ’거래량’, ‘거래대금’, ‘날짜’, ‘시가’, ‘고가’, ‘저가’, ‘’]. […]]

            
            cnt = self.dynamicCall("GetRepeatCnt(QString, QString)", sTrCode, sRQName)
            print("전송받는 data 총 일수 : %s " % cnt)
            
            #한번조회하면 600일치까지 일봉데이터를 받을 수 있음.
            
            for i in range(cnt):
                data = []
                
                current_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "현재가")  # 출력 : 000070
                value = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "거래량")  # 출력 : 000070
                trading_value = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "거래대금")  # 출력 : 000070
                date = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "일자")  # 출력 : 000070
                start_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "시가")  # 출력 : 000070
                high_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "고가")  # 출력 : 000070
                low_price = self.dynamicCall("GetCommData(QString, QString, int, QString)", sTrCode, sRQName, i, "저가")  # 출력 : 000070
                ###############[check]좀 더 효율적으로 종목 List를 불러올 수 없을지 생각해봐야함.##############
                data.append("")
                data.append(current_price.strip())
                data.append(value.strip())
                data.append(trading_value.strip())
                data.append(date.strip())
                data.append(start_price.strip())
                data.append(high_price.strip())
                data.append(low_price.strip())
                data.append("")
                self.calcul_data.append(data.copy())
                
            print(len(self.calcul_data))

            if sPrevNext == "2" :  # 데이터가 더 있으면 한번 더 조회신청하게됨
                self.day_kiwoom_db(code=code, sPrevNext=sPrevNext)
                
            else : 
                self.calculator_event_loop.exit()
            

